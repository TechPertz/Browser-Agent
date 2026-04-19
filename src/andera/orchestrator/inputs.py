"""Input file loader — CSV / JSONL / Excel (xlsx via openpyxl if installed).

Returns a list of dicts; downstream code is format-agnostic.
"""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any


def load_inputs(path: str | Path) -> list[dict[str, Any]]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"input file not found: {p}")
    suffix = p.suffix.lower()
    if suffix == ".csv":
        return _load_csv(p)
    if suffix in (".jsonl", ".ndjson"):
        return _load_jsonl(p)
    if suffix == ".json":
        return _load_json(p)
    if suffix in (".xlsx", ".xls"):
        return _load_xlsx(p)
    raise ValueError(f"unsupported input format: {suffix}")


def _load_csv(p: Path) -> list[dict[str, Any]]:
    with p.open(newline="") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def _load_jsonl(p: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with p.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def _load_json(p: Path) -> list[dict[str, Any]]:
    data = json.loads(p.read_text())
    if isinstance(data, list):
        return data
    raise ValueError("JSON input must be a list of objects")


def _load_xlsx(p: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "xlsx inputs require openpyxl. Add it with `uv add openpyxl`."
        ) from e
    wb = load_workbook(p, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row, strict=False)) for row in rows[1:]]
